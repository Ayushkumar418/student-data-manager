# Global imports and Flask app setup
from flask import Flask, render_template, request, redirect, url_for, jsonify, send_file, flash
import json
import os
from io import StringIO, BytesIO  # Add this import at the top
from datetime import datetime
from openpyxl import Workbook, load_workbook

app = Flask(__name__)
app.secret_key = 'your_secret_key'  # Add a secret key for flash messages

# Student data management functions and global variable
DATA_FILE = "students_data.json"  # Constants defined at top for easy configuration

def load_students():
    """
    Load student data from JSON file.
    Returns empty list if file doesn't exist or is corrupted.
    """
    try:
        if os.path.exists(DATA_FILE) and os.path.getsize(DATA_FILE) > 0:
            with open(DATA_FILE, "r") as file:
                return json.load(file)
    except json.JSONDecodeError:
        print("Error reading JSON file. Creating new file.")
    except Exception as e:
        print(f"Error loading students: {str(e)}")
    return []

def save_students(students):
    """
    Save student data to JSON file with pretty printing.
    Args:
        students: List of student dictionaries to save
    """
    with open(DATA_FILE, "w") as file:
        json.dump(students, file, indent=4)

# Global variable to store students in memory
students = load_students()

@app.route("/")
def index():
    """Main entry point - renders the home page"""
    return render_template("index.html")

@app.route("/add_student", methods=["POST"])
def add_student():
    try:
        # Basic student fields with validation
        new_student = {
            "name": request.form.get("name", "").strip(),
            "course": request.form.get("course", "").strip(),
            "roll": request.form.get("roll", "").strip(),
            "sec": request.form.get("sec", "").strip().upper(),
            "year": request.form.get("year", "").strip(),
            "cgpa": float(request.form.get("cgpa", 0)),
            "additional": {}
        }

        # Handle additional fields
        for key in request.form:
            if key.startswith('additional_field_'):
                value = request.form.get(key).strip()
                if value:  # Only add if value is not empty
                    new_student["additional"][key] = value

        # Validate required fields
        if not all([new_student["name"], new_student["course"], 
                   new_student["roll"], new_student["sec"], 
                   new_student["year"]]):  # Fixed syntax error here
            flash('All fields are required', 'error')
            return redirect(url_for("index"))

        students.append(new_student)
        save_students(students)
        flash('Student added successfully', 'success')
        return redirect(url_for("view_students"))  # Redirect to view page instead
    except Exception as e:
        print(f"Error adding student: {str(e)}")
        flash('Error adding student', 'error')
        return redirect(url_for("index"))

@app.route("/view_students")
def view_students():
    return render_template("view_students.html", students=students)

@app.route("/edit_student/<int:index>", methods=["GET", "POST"])
def edit_student(index):
    if request.method == "POST":
        # Update basic student data
        students[index]["name"] = request.form.get("name")
        students[index]["course"] = request.form.get("course")
        students[index]["roll"] = request.form.get("roll")
        students[index]["sec"] = request.form.get("sec")
        students[index]["year"] = request.form.get("year")
        students[index]["cgpa"] = request.form.get("cgpa")
        
        # Update additional fields
        students[index]["additional"] = {}
        for key in request.form:
            if key.startswith("additional_field_"):
                students[index]["additional"][key] = request.form[key]

        save_students(students)  # Save updated data
        flash('Student updated successfully', 'success')
        return redirect(url_for("view_students"))
    
    return render_template("edit_student.html", student=students[index], index=index)

@app.route("/delete_student/<int:index>")
def delete_student(index):
    del students[index]
    save_students(students)  # Save updated data
    flash('Student deleted successfully', 'success')
    return redirect(url_for("view_students"))

@app.route("/search")
def search_students():
    query = request.args.get('query', '')
    field = request.args.get('field', 'name')
    
    filtered_students = [
        student for student in students
        if str(student.get(field, '')).lower().startswith(query.lower())
    ]
    return jsonify(filtered_students)

@app.route("/statistics")
def get_statistics():
    try:
        if not students:
            return render_template("statistics.html", stats={
                "total_students": 0,
                "average_cgpa": 0,
                "courses": {},
                "years": {}
            })

        stats = {
            "total_students": len(students),
            "average_cgpa": sum(float(s.get('cgpa', 0)) for s in students) / len(students),
            "courses": {},
            "years": {}
        }
        
        for student in students:
            course = student.get('course', 'Unknown')
            year = student.get('year', 'Unknown')
            stats['courses'][course] = stats['courses'].get(course, 0) + 1
            stats['years'][year] = stats['years'].get(year, 0) + 1
        
        return render_template("statistics.html", stats=stats)
    except Exception as e:
        print(f"Error generating statistics: {str(e)}")
        flash("Error generating statistics", "error")
        return redirect(url_for("view_students"))

@app.route("/export")
def export_students():
    try:
        output = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Students Data"
        
        # Get all possible additional field keys
        additional_fields = set()
        for student in students:
            additional_fields.update(student['additional'].keys())
        
        # Write headers including additional fields
        headers = ['name', 'course', 'roll', 'sec', 'year', 'cgpa'] + list(additional_fields)
        ws.append(headers)
        
        # Write student data
        for student in students:
            row = [student.get(field, '') for field in headers[:6]]  # Basic fields
            # Add additional fields
            for field in additional_fields:
                row.append(student['additional'].get(field, ''))
            ws.append(row)

        # Save to BytesIO
        wb.save(output)
        output.seek(0)
        
        filename = f'students_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        print(f"Export error: {str(e)}")
        flash('Error exporting data', 'error')
        return redirect(url_for('view_students'))

@app.route("/download_template")
def download_template():
    try:
        output = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Template"
        
        # Add headers
        headers = ['name', 'course', 'roll', 'sec', 'year', 'cgpa']
        ws.append(headers)
        
        # Add sample row
        sample_row = ['John Doe', 'B.Tech', 'BT123', 'A', '1', '8.5']
        ws.append(sample_row)
        
        # Add column descriptions in the next sheet
        ws2 = wb.create_sheet("Instructions")
        ws2.append(['Column', 'Description', 'Format', 'Required'])
        ws2.append(['name', 'Student Name', 'Text', 'Yes'])
        ws2.append(['course', 'Course Name', 'Text (B.Tech, M.Tech, etc)', 'Yes'])
        ws2.append(['roll', 'Roll Number', 'Text', 'Yes'])
        ws2.append(['sec', 'Section', 'Single Letter (A-Z)', 'Yes'])
        ws2.append(['year', 'Year', 'Number (1-6)', 'Yes'])
        ws2.append(['cgpa', 'CGPA', 'Number (0-10)', 'Yes'])

        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='student_data_template.xlsx'
        )
    except Exception as e:
        flash('Error generating template', 'error')
        return redirect(url_for('view_students'))

@app.route("/import", methods=["POST"])
def import_students():
    if 'file' not in request.files:
        flash('Please select a file to import', 'error')
        return redirect(url_for('view_students'))

    file = request.files['file']
    if not file or file.filename == '':
        flash('No file selected', 'error')
        return redirect(url_for('view_students'))

    # Check MIME type and extension
    allowed_mimetypes = ['application/vnd.ms-excel',
                        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet']
    if file.mimetype not in allowed_mimetypes:
        flash('Invalid file type! Please upload only Excel files.', 'error')
        return redirect(url_for('view_students'))

    # Validate file extension
    file_ext = os.path.splitext(file.filename)[1].lower()
    if file_ext not in ['.xlsx', '.xls']:
        flash('Invalid file extension! Please upload .xlsx or .xls files only.', 'error')
        return redirect(url_for('view_students'))

    try:
        # Try to read file content
        file_content = file.read()
        
        # Try to load as Excel
        try:
            wb = load_workbook(BytesIO(file_content))
        except Exception as e:
            flash('Invalid Excel format! Please use the correct template format.', 'error')
            return redirect(url_for('view_students'))

        ws = wb.active
        if not ws or ws.max_row < 2:
            flash('Excel file is empty or invalid! Please use the template.', 'error')
            return redirect(url_for('view_students'))

        # Validate headers (first row)
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value).lower().strip())

        required_headers = ['name', 'course', 'roll', 'sec', 'year', 'cgpa']
        missing = [h for h in required_headers if h not in headers]
        if missing:
            raise ValueError(f'Missing required columns: {", ".join(missing)}')

        # Process data rows
        imported_students = []
        error_rows = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if not any(cell.value for cell in row):
                continue

            try:
                # Create student data dictionary
                data = {headers[i]: (str(cell.value).strip() if cell.value is not None else '')
                       for i, cell in enumerate(row) if i < len(headers)}

                # Validate required fields
                if not all(data.get(field) for field in required_headers[:-1]):  # Excluding CGPA
                    raise ValueError('Missing required fields')

                # Validate CGPA
                try:
                    cgpa = float(data.get('cgpa', '0'))
                    if not 0 <= cgpa <= 10:
                        raise ValueError('CGPA must be between 0 and 10')
                except ValueError:
                    raise ValueError('Invalid CGPA value')

                # Create validated student object
                new_student = {
                    "name": data['name'],
                    "course": data['course'],
                    "roll": data['roll'],
                    "sec": data['sec'].upper(),
                    "year": data['year'],
                    "cgpa": cgpa,
                    "additional": {
                        f"additional_field_{header}": data[header]
                        for header in headers
                        if header not in required_headers and data.get(header)
                    }
                }
                imported_students.append(new_student)

            except Exception as e:
                error_rows.append(f"Row {row_idx}: {str(e)}")

        if not imported_students and error_rows:
            raise ValueError("No valid data found. Please check the file format")

        # Save valid data
        if imported_students:
            students.extend(imported_students)
            save_students(students)
            flash(f'Successfully imported {len(imported_students)} students', 'success')
            if error_rows:
                for error in error_rows:
                    flash(error, 'warning')

    except ValueError as ve:
        flash(f'Import Error: {str(ve)}', 'error')
    except Exception as e:
        flash('Error: Invalid file format. Please use the template provided.', 'error')
    
    return redirect(url_for('view_students'))

if __name__ == "__main__":
    app.run(debug=True)